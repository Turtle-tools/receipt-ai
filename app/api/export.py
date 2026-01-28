"""
Export endpoints for downloading processed data.
"""

import csv
import json
import io
from datetime import datetime, timedelta
from typing import Optional, List
from fastapi import APIRouter, Depends, Query, Response, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.core.database import get_db
from app.core.auth import get_api_key
from app.models.database import Document

router = APIRouter()


class ExportFilter(BaseModel):
    """Filters for export."""
    company_id: Optional[str] = None
    document_type: Optional[str] = None
    status: Optional[str] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None


@router.get("/csv")
async def export_csv(
    company_id: Optional[str] = None,
    status: Optional[str] = None,
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    api_key: str = Depends(get_api_key),
):
    """
    Export extracted document data as CSV.
    
    Includes all extracted fields in a flat format.
    Great for importing into spreadsheets or other accounting software.
    """
    # Build query
    query = db.query(Document)
    
    if company_id:
        query = query.filter(Document.company_id == company_id)
    
    if status:
        query = query.filter(Document.status == status)
    
    # Filter by date
    start_date = datetime.utcnow() - timedelta(days=days)
    query = query.filter(Document.created_at >= start_date)
    
    # Order by date
    query = query.order_by(Document.created_at.desc())
    
    documents = query.all()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "ID",
        "Filename",
        "Upload Date",
        "Status",
        "Document Type",
        "Vendor",
        "Amount",
        "Date",
        "Category",
        "QBO ID",
        "QBO Status",
    ])
    
    # Data rows
    for doc in documents:
        extracted = doc.extracted_data or {}
        
        writer.writerow([
            doc.id,
            doc.filename,
            doc.created_at.isoformat(),
            doc.status,
            doc.document_type or "",
            extracted.get("vendor", ""),
            extracted.get("total_amount", ""),
            extracted.get("date", ""),
            extracted.get("category_suggestion", ""),
            doc.qbo_id or "",
            doc.qbo_sync_status or "",
        ])
    
    # Return as download
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=receipt-ai-export-{datetime.now().strftime('%Y%m%d')}.csv"
        },
    )


@router.get("/json")
async def export_json(
    company_id: Optional[str] = None,
    status: Optional[str] = None,
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    api_key: str = Depends(get_api_key),
):
    """
    Export extracted document data as JSON.
    
    Includes full nested data structure.
    Great for programmatic processing or backup.
    """
    # Build query
    query = db.query(Document)
    
    if company_id:
        query = query.filter(Document.company_id == company_id)
    
    if status:
        query = query.filter(Document.status == status)
    
    # Filter by date
    start_date = datetime.utcnow() - timedelta(days=days)
    query = query.filter(Document.created_at >= start_date)
    
    # Order by date
    query = query.order_by(Document.created_at.desc())
    
    documents = query.all()
    
    # Serialize
    data = {
        "exported_at": datetime.utcnow().isoformat(),
        "total_documents": len(documents),
        "filters": {
            "company_id": company_id,
            "status": status,
            "days": days,
        },
        "documents": [
            {
                "id": doc.id,
                "filename": doc.filename,
                "file_type": doc.file_type,
                "document_type": doc.document_type,
                "status": doc.status,
                "created_at": doc.created_at.isoformat() if doc.created_at else None,
                "updated_at": doc.updated_at.isoformat() if doc.updated_at else None,
                "extracted_data": doc.extracted_data,
                "qbo_id": doc.qbo_id,
                "qbo_sync_status": doc.qbo_sync_status,
                "metadata": doc.metadata,
            }
            for doc in documents
        ],
    }
    
    # Return as download
    json_str = json.dumps(data, indent=2)
    
    return Response(
        content=json_str,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=receipt-ai-export-{datetime.now().strftime('%Y%m%d')}.json"
        },
    )


@router.get("/qbo-import")
async def export_qbo_import_format(
    company_id: Optional[str] = None,
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    api_key: str = Depends(get_api_key),
):
    """
    Export in QuickBooks import format (IIF/CSV).
    
    Creates a CSV that can be imported directly into QuickBooks
    using their import feature.
    """
    # Build query for extracted receipts
    query = db.query(Document).filter(
        Document.status == "extracted",
        Document.document_type == "receipt",
    )
    
    if company_id:
        query = query.filter(Document.company_id == company_id)
    
    # Filter by date
    start_date = datetime.utcnow() - timedelta(days=days)
    query = query.filter(Document.created_at >= start_date)
    
    documents = query.all()
    
    # Create QuickBooks-compatible CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header (QuickBooks format)
    writer.writerow([
        "!TRNS",
        "TRNSID",
        "TRNSTYPE",
        "DATE",
        "ACCNT",
        "NAME",
        "AMOUNT",
        "DOCNUM",
        "MEMO",
        "CLEAR",
    ])
    
    # Split rows
    writer.writerow([
        "!SPL",
        "SPLID",
        "TRNSTYPE",
        "DATE",
        "ACCNT",
        "NAME",
        "AMOUNT",
        "DOCNUM",
        "MEMO",
        "CLEAR",
    ])
    
    # Data rows
    for doc in documents:
        extracted = doc.extracted_data or {}
        
        vendor = extracted.get("vendor", "Unknown Vendor")
        amount = extracted.get("total_amount", 0)
        date = extracted.get("date", datetime.now().strftime("%m/%d/%Y"))
        category = extracted.get("category_suggestion", "Expenses")
        
        # Transaction row
        writer.writerow([
            "TRNS",
            "",
            "CREDIT CARD",
            date,
            "Credit Card",
            vendor,
            f"-{amount}",
            doc.filename,
            extracted.get("notes", ""),
            "N",
        ])
        
        # Split row (expense account)
        writer.writerow([
            "SPL",
            "",
            "CREDIT CARD",
            date,
            category,
            vendor,
            amount,
            doc.filename,
            "",
            "N",
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=qbo-import-{datetime.now().strftime('%Y%m%d')}.csv"
        },
    )


@router.get("/excel")
async def export_excel(
    company_id: Optional[str] = None,
    status: Optional[str] = None,
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    api_key: str = Depends(get_api_key),
):
    """
    Export as Excel file (.xlsx).
    
    Requires openpyxl: pip install openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="Excel export requires openpyxl. Install with: pip install openpyxl"
        )
    
    # Build query
    query = db.query(Document)
    
    if company_id:
        query = query.filter(Document.company_id == company_id)
    
    if status:
        query = query.filter(Document.status == status)
    
    # Filter by date
    start_date = datetime.utcnow() - timedelta(days=days)
    query = query.filter(Document.created_at >= start_date)
    
    documents = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"
    
    # Header styling
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = [
        "ID", "Filename", "Upload Date", "Status", "Document Type",
        "Vendor", "Amount", "Date", "Category", "QBO ID", "QBO Status"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row, doc in enumerate(documents, 2):
        extracted = doc.extracted_data or {}
        
        ws.cell(row=row, column=1, value=doc.id)
        ws.cell(row=row, column=2, value=doc.filename)
        ws.cell(row=row, column=3, value=doc.created_at.strftime("%Y-%m-%d %H:%M") if doc.created_at else "")
        ws.cell(row=row, column=4, value=doc.status)
        ws.cell(row=row, column=5, value=doc.document_type or "")
        ws.cell(row=row, column=6, value=extracted.get("vendor", ""))
        ws.cell(row=row, column=7, value=extracted.get("total_amount", ""))
        ws.cell(row=row, column=8, value=extracted.get("date", ""))
        ws.cell(row=row, column=9, value=extracted.get("category_suggestion", ""))
        ws.cell(row=row, column=10, value=doc.qbo_id or "")
        ws.cell(row=row, column=11, value=doc.qbo_sync_status or "")
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=receipt-ai-export-{datetime.now().strftime('%Y%m%d')}.xlsx"
        },
    )
